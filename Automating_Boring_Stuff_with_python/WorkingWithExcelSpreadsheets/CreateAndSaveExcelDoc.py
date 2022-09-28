# Creating and saving Excel file.

import openpyxl

# ------- Create -------
wb = openpyxl.Workbook()        # Create a blank workbook.
print(wb.sheetnames)        # It start with one sheet
sheet = wb.active       # Loading sheet
print(sheet.title)      # Get sheet title
sheet.title = 'Spam Bacon Eggs Sheet'       # Change title
print(wb.sheetnames)        # Get new changed title

# ------- Save --------
wb = openpyxl.load_workbook('example.xlsx')     # Open existing file
sheet = wb.active       # Loading sheet
sheet.title = 'Spam Spam Spam'      # Changing title
wb.save('example_copy_to_del.xlsx')        # Save the notebook to new file
