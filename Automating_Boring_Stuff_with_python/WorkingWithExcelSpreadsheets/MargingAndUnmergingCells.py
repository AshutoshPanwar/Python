import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')      # Merge all these cells.
sheet['A1'] = 'Twelve cells merged togeather.'
sheet.merge_cells('C5:D5')      # Merge all these cells.
sheet['C5'] = 'Two merged cells.'
print("New file is created with name: merged.xlsx")
wb.save('merged.xlsx')
