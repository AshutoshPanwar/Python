import openpyxl

wb = openpyxl.Workbook()        # New filea

# ------- Creating Sheet -------
print(wb.sheetnames)
print("Adding Sheet's :")
wb.create_sheet()       # Add a new sheet.
print(wb.sheetnames)
# Create a new sheet at index 0.
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)

# ------- Deleting Sheet -------
print("Delet sheet's :")
del wb['Middle Sheet']
del wb['Sheet1']
print(wb.sheetnames)